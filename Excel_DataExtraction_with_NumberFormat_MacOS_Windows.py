# -*- coding: utf-8 -*-
import openpyxl
import xlwings
import pandas as pd
import re
# ================================================FUNCTIONS================================================
# a function that convert the excel-uncaptured dates into DD/MM/YYYY
def change_date(match_obj):
  day = match_obj.group('day')
  month = match_obj.group('month')
  year = match_obj.group('year')

  if len(day) == 1:
    day = '0' + day
  
  if len(month) == 1:
    month = '0' + month
  
  elif len(month) == 3:
    month_dict = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04','May': '05', 'Jun': '06',
                  'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10','Nov': '11', 'Dec': '12'}
    month = month_dict.get(month) 
  
  if len(year) == 2:
    year = '20' + year

  return f'{day}/{month}/{year}'
  
  
# a function that convert number format into an excel formula
def apply_number_format(cell):  
  if cell.value is not None:
    
    if cell.data_type == 'd':
      cell.value = f'=TEXT("{cell.value}", "dd/mm/yyyy")'
    
    elif cell.data_type == 's':
      # regex for capturing [dd.m.yy] [dd.mm.yy] [dd.m.yyyy] [dd.mm.yyyy]
      date_pattern1 = r"(?:(?P<day>\d{1,2})\.(?P<month>\d{1,2})\.(?P<year>\d{2}(?:\d{2})?))"
      # regex for capturing [month dd yyyy] [month ddth yyyy] [month dd, yyyy] [month ddth, yyyy]
      date_pattern2 = r"(?P<month>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]* (?:(?P<day>\d{1,2})(?:st|nd|rd|th)?,? (?P<year>\d{4}))"
      # regex for capturing [dd month yyyy] [ddth month yyyy] [dd month, yyyy] [ddth month, yyyy]
      date_pattern3 = r"(?:(?P<day>\d{1,2})(?:st|nd|rd|th)? (?P<month>Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*,? (?P<year>\d{4}))"

      unconverted_date = re.match(date_pattern1, cell.value) or re.match(date_pattern2, cell.value) or re.match(date_pattern3, cell.value)
    
      if unconverted_date:
        cell.value = change_date(unconverted_date)
      elif cell.number_format != 'General':
        cell.value = f'=IFERROR(TEXT("{cell.value}", "{cell.number_format}"), "{cell.value}")'
    
    elif cell.number_format != 'General':
      cell.value = f'=IFERROR(TEXT("{cell.value}", "{cell.number_format}"), "{cell.value}")'

  cell.number_format = 'General'


# ==============================================MAIN PROGRAM===============================================
# path to the targeted file
path = '/path/to/directory/example.xlsx'

# flatten existing equations
wb = openpyxl.load_workbook(path, data_only=True)

# for every cell in the workbook, call the apply_number_format() function
for ws in wb:
  # iterate the merged cells
  for merged_ranges in ws.merged_cells:
    min_col, min_row, max_col, max_row = merged_ranges.bounds
    top_left_cell = ws.cell(row=min_row, column=min_col)
    top_left_cell = apply_number_format(top_left_cell)

  # iterate the unmerged cells
  for row in ws.iter_rows(min_row=ws.min_row, min_col=ws.min_column, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
      cell = apply_number_format(cell)

# save the modified workbook
wb.save(path)

# open the excel file in the background with xlwings such that the formulas' value are computed by excel
excel_app = xlwings.App(visible=False)
excel_book = excel_app.books.open(path)
excel_book.save()
excel_book.close()
excel_app.quit()

# flatten all formulas
wb = openpyxl.load_workbook(path, data_only=True)
wb.save(path)

# store the excel file in pandas 
excel_file = pd.ExcelFile(path)

# Create an empty dictionary to store the DataFrames for each sheet
dfs = {}

# Iterate over each sheet name in the Excel file
for sheet_name in excel_file.sheet_names:
    # Read the sheet into a DataFrame with correct indexing and titles
    df = pd.read_excel(excel_file, sheet_name=sheet_name, index_col=None)
    # Store the DataFrame in the dictionary with the sheet name as the key
    dfs[sheet_name] = df

# integrate with other programs and get the result from dfs
print(dfs["DETAIL"])